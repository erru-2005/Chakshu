from flask import Flask, render_template, request, jsonify, url_for, flash, redirect, send_file, session
import os
from dotenv import load_dotenv
import json
import pandas as pd
from werkzeug.utils import secure_filename
from PIL import Image
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side
from functools import wraps
from db import generate_otp, save_otp, verify_otp, get_student_mobile
from twilio.rest import Client
from flask import current_app

# Load environment variables
load_dotenv()

app = Flask(__name__)

# Add secret key for flash messages and session
app.secret_key = os.environ.get('SECRET_KEY', 'default_secret_key_for_development')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 86400  # cache static files for a day

# Configure upload folder
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# JSON storage file path
STUDENTS_JSON = 'students.json'

def load_students():
    if os.path.exists(STUDENTS_JSON):
        with open(STUDENTS_JSON, 'r') as f:
            return json.load(f)
    return []

def save_students(students):
    with open(STUDENTS_JSON, 'w') as f:
        json.dump(students, f, indent=4)

def _resolve_profile_image_internal(profile_image, roll_no):
    """Return relative static path for an existing profile image.
    Tries stored path first, then guesses by roll number with common extensions and cases.
    """
    try:
        static_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
        candidates = []
        if profile_image:
            candidates.append(os.path.join(static_root, str(profile_image).replace('\\', '/')))
        roll_str = (str(roll_no or '').strip())
        if roll_str:
            for ext in ('jpg', 'jpeg', 'png'):
                candidates.append(os.path.join(static_root, 'uploads', f"{roll_str.lower()}.{ext}"))
                candidates.append(os.path.join(static_root, 'uploads', f"{roll_str.upper()}.{ext}"))
        for abs_path in candidates:
            if os.path.exists(abs_path):
                return os.path.relpath(abs_path, static_root).replace('\\', '/')
    except Exception:
        pass
    return None

def _resolve_profile_image_thumb(profile_image, roll_no):
    try:
        static_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')
        roll_str = (str(roll_no or '').strip())
        thumb_webp = os.path.join(static_root, 'uploads', 'thumbs', f"{roll_str.lower()}_thumb.webp")
        if os.path.exists(thumb_webp):
            return os.path.relpath(thumb_webp, static_root).replace('\\', '/')
    except Exception:
        pass
    return _resolve_profile_image_internal(profile_image, roll_no)

@app.context_processor
def inject_image_resolver():
    return {
        'resolve_profile_image': _resolve_profile_image_internal,
        'resolve_profile_image_thumb': _resolve_profile_image_thumb
    }

def _avatar_style(name):
    try:
        palette = [
            'linear-gradient(135deg, hsla(210, 85%, 40%, 0.90), hsla(210, 90%, 55%, 0.75))',
            'linear-gradient(135deg, hsla(260, 70%, 45%, 0.90), hsla(260, 80%, 60%, 0.75))',
            'linear-gradient(135deg, hsla(340, 75%, 45%, 0.90), hsla(340, 85%, 60%, 0.75))',
            'linear-gradient(135deg, hsla(28, 85%, 50%, 0.90), hsla(28, 95%, 60%, 0.75))',
            'linear-gradient(135deg, hsla(140, 55%, 40%, 0.90), hsla(140, 65%, 50%, 0.75))',
            'linear-gradient(135deg, hsla(190, 65%, 40%, 0.90), hsla(190, 75%, 55%, 0.75))',
            'linear-gradient(135deg, hsla(50, 85%, 45%, 0.90), hsla(50, 95%, 55%, 0.75))',
            'linear-gradient(135deg, hsla(280, 70%, 45%, 0.90), hsla(280, 80%, 60%, 0.75))',
        ]
        s = str(name or '').lower()
        if not s:
            return palette[0]
        h = 0
        for ch in s:
            h = ((h << 5) - h) + ord(ch)
            h &= 0xFFFFFFFF
        return palette[abs(h) % len(palette)]
    except Exception:
        return 'linear-gradient(135deg, hsla(210, 85%, 40%, 0.90), hsla(210, 90%, 55%, 0.75))'

def _avatar_theme(name):
    palettes = [
        ("linear-gradient(135deg, hsla(210, 85%, 40%, 0.90), hsla(210, 90%, 55%, 0.75))", "#ffffff"),
        ("linear-gradient(135deg, hsla(260, 70%, 45%, 0.90), hsla(260, 80%, 60%, 0.75))", "#ffffff"),
        ("linear-gradient(135deg, hsla(340, 75%, 45%, 0.90), hsla(340, 85%, 60%, 0.75))", "#ffffff"),
        ("linear-gradient(135deg, hsla(28, 85%, 50%, 0.90), hsla(28, 95%, 60%, 0.75))", "#1f2937"),
        ("linear-gradient(135deg, hsla(140, 55%, 40%, 0.90), hsla(140, 65%, 50%, 0.75))", "#ffffff"),
        ("linear-gradient(135deg, hsla(190, 65%, 40%, 0.90), hsla(190, 75%, 55%, 0.75))", "#ffffff"),
        ("linear-gradient(135deg, hsla(50, 85%, 45%, 0.90), hsla(50, 95%, 55%, 0.75))", "#1f2937"),
        ("linear-gradient(135deg, hsla(280, 70%, 45%, 0.90), hsla(280, 80%, 60%, 0.75))", "#ffffff"),
    ]
    s = str(name or '').lower()
    if not s:
        return { 'bg': palettes[0][0], 'fg': palettes[0][1] }
    h = 0
    for ch in s:
        h = ((h << 5) - h) + ord(ch)
        h &= 0xFFFFFFFF
    bg, fg = palettes[abs(h) % len(palettes)]
    return { 'bg': bg, 'fg': fg }

@app.context_processor
def inject_avatar_style():
    return {
        'avatar_style': _avatar_style,
        'avatar_theme': _avatar_theme
    }

def login_required(f):
    """Decorator to check if user is logged in"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session or not session['logged_in']:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def admin_login_required(f):
    """Decorator to protect admin routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login_page'))
        return f(*args, **kwargs)
    return decorated_function

# Login Routes
@app.route('/login')
def login_page():
    """Render the login page"""
    if 'logged_in' in session and session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    """Verify roll number and DOB"""
    try:
        data = request.json
        roll_no = data.get('rollNo', '').strip()
        dob = data.get('dob', '').strip()
        
        if not roll_no or not dob:
            return jsonify({'success': False, 'error': 'Roll number and DOB are required'}), 400
        
        # Load students from JSON
        students = load_students()
        student = next((s for s in students if s.get('rollNo', '').upper() == roll_no.upper()), None)
        
        if not student:
            return jsonify({'success': False, 'error': 'Invalid roll number'}), 404
        
        # Compare DOB - normalize both formats
        student_dob = str(student.get('dob', '')).strip()
        
        # Try different date formats
        dob_normalized = dob.replace('/', '-').replace('.', '-')
        student_dob_normalized = student_dob.replace('/', '-').replace('.', '-')
        
        if dob_normalized.lower() != student_dob_normalized.lower():
            return jsonify({'success': False, 'error': 'Invalid date of birth'}), 401
        
        # Store roll number in session for next steps
        session['roll_no'] = roll_no.upper()
        session['student_data'] = student
        
        return jsonify({'success': True, 'message': 'Credentials verified'}), 200
        
    except Exception as e:
        print(f"Error in api_login: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/send-otp', methods=['POST'])
def api_send_otp():
    """Send OTP to the already verified mobile number for the roll number"""
    try:
        data = request.json
        roll_no = (data.get('rollNo') or '').strip()

        if not roll_no:
            return jsonify({'success': False, 'error': 'Roll number is required'}), 400

        # Check if roll_no matches session
        if 'roll_no' not in session or session['roll_no'].upper() != roll_no.upper():
            return jsonify({'success': False, 'error': 'Session expired. Please login again.'}), 401

        # Fetch verified mobile from MongoDB, otherwise fall back to studentContact from JSON
        mobile = get_student_mobile(roll_no.upper())
        if not mobile:
            # Fallback: read studentContact from students.json
            students = load_students()
            student = next((s for s in students if s.get('rollNo', '').upper() == roll_no.upper()), None)
            if student:
                fallback = str(student.get('studentContact') or '').strip()
                # Normalize fallback number by removing spaces and hyphens
                fallback_clean = ''.join(c for c in fallback if c.isdigit())
                # Use only if it looks like an Indian 10-digit number
                if len(fallback_clean) == 10 and fallback_clean[0] in ['6','7','8','9']:
                    mobile = fallback_clean
            
        if not mobile:
            return jsonify({'success': False, 'error': 'No verified mobile found for this roll number.'}), 403

        # Generate and save OTP
        otp = generate_otp(6)
        save_otp(mobile, roll_no.upper(), otp)

        # Store mobile in session
        session['mobile'] = mobile

        # Prepare masked mobile for display (e.g., +91 XXXXXXX123)
        # Build masked number based on last 3 digits of the cleaned number
        last_three = ''.join(c for c in str(mobile) if c.isdigit())[-3:]
        masked = f"+91 {'X'*7}{last_three}" if last_three else "+91 **********"

        # Send OTP via Twilio SMS
        try:
            # Get Twilio credentials from environment variables
            account_sid = os.environ.get('account_sid')
            auth_token = os.environ.get('auth_token')
            twilio_number = os.environ.get('twilio_number')
            
            if not all([account_sid, auth_token, twilio_number]):
                print("Warning: Twilio credentials not found in .env file. OTP will not be sent via SMS.")
                print(f"OTP for {roll_no} to {mobile}: {otp}")  # Fallback: Print OTP for testing
            else:
                # Format mobile number (ensure it starts with +91)
                # Remove all spaces, dashes, and other non-digit characters (except +)
                recipient_number = ''.join(c for c in mobile.strip() if c.isdigit() or c == '+')
                
                if not recipient_number.startswith('+91'):
                    # Remove leading + if present for processing
                    clean_number = recipient_number.lstrip('+')
                    
                    # Remove leading 91 if present (India country code)
                    if clean_number.startswith('91') and len(clean_number) > 10:
                        clean_number = clean_number[2:]
                    
                    # Remove leading 0 if present
                    if clean_number.startswith('0'):
                        clean_number = clean_number[1:]
                    
                    # Ensure it's a 10-digit number, then add +91
                    if len(clean_number) == 10 and clean_number.isdigit():
                        recipient_number = '+91' + clean_number
                    else:
                        raise ValueError(f"Invalid mobile number format: {mobile}")
                
                # Create Twilio client
                client = Client(account_sid, auth_token)
                
                # Send the OTP message
                message = client.messages.create(
                    body=f"Your OTP for CHAKSHU is: {otp}. Valid for 4 minutes. Do not share this OTP with anyone.Regards: Incubation Center",
                    from_=twilio_number,
                    to=recipient_number
                )
                
                print(f"OTP SMS sent successfully to {recipient_number}. Message SID: {message.sid}")
        
        except Exception as sms_error:
            # If SMS sending fails, log the error but don't fail the request
            print(f"Failed to send SMS via Twilio: {str(sms_error)}")
            print(f"OTP for {roll_no} to {mobile}: {otp}")  # Fallback: Print OTP for debugging
            
        return jsonify({
            'success': True,
            'message': 'OTP sent successfully',
            'maskedMobile': masked
        }), 200

    except Exception as e:
        print(f"Error in api_send_otp: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/verify-otp', methods=['POST'])
def api_verify_otp():
    """Verify OTP and complete login"""
    try:
        data = request.json
        roll_no = data.get('rollNo', '').strip()
        entered_otp = data.get('otp', '').strip()
        
        if not roll_no or not entered_otp:
            return jsonify({'success': False, 'error': 'All fields are required'}), 400
        
        if len(entered_otp) != 6:
            return jsonify({'success': False, 'error': 'OTP must be 6 digits'}), 400
        
        # Determine mobile (from session or DB) and verify OTP using MongoDB
        mobile = session.get('mobile') or get_student_mobile(roll_no.upper())
        if not mobile:
            return jsonify({'success': False, 'error': 'Verified mobile not found'}), 403

        if verify_otp(mobile, roll_no.upper(), entered_otp):
            # Load student data
            students = load_students()
            student = next((s for s in students if s.get('rollNo', '').upper() == roll_no.upper()), None)
            
            if student:
                # Set session
                session['logged_in'] = True
                session['roll_no'] = roll_no.upper()
                session['mobile'] = mobile
                session['student_data'] = student
                
                return jsonify({'success': True, 'message': 'Login successful'}), 200
            else:
                return jsonify({'success': False, 'error': 'Student not found'}), 404
        else:
            return jsonify({'success': False, 'error': 'Invalid OTP'}), 401
        
    except Exception as e:
        print(f"Error in api_verify_otp: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    return redirect(url_for('login_page'))

# Admin Auth Routes
@app.route('/admin/login', methods=['GET'])
def admin_login_page():
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_submit():
    try:
        username = (request.form.get('username') or '').strip()
        password = (request.form.get('password') or '').strip()

        if username == 'admin' and password == 'admin@2005':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials', 'error')
            return redirect(url_for('admin_login_page'))
    except Exception as e:
        flash(str(e), 'error')
        return redirect(url_for('admin_login_page'))

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin_login_page'))

@app.route('/')
@login_required
def dashboard():
    """Student dashboard/home page"""
    try:
        roll_no = session.get('roll_no')
        if not roll_no:
            return redirect(url_for('login_page'))
        
        # Load student data from JSON
        students = load_students()
        student = next((s for s in students if s.get('rollNo', '').upper() == roll_no.upper()), None)
        
        if not student:
            session.clear()
            return redirect(url_for('login_page'))
        
        return render_template('dashboard.html', student=student)
        
    except Exception as e:
        print(f"Error in dashboard: {str(e)}")
        session.clear()
        return redirect(url_for('login_page'))

@app.route('/admin')
def admin_root_redirect():
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/dashboard')
@admin_login_required
def admin_dashboard():
    return render_template('student_form.html')

@app.route('/search_barcode', methods=['POST'])
def search_barcode():
    try:
        data = request.json
        barcode = data.get('barcode')
        
        if not barcode:
            return jsonify({'error': 'Barcode is required'}), 400

        students = load_students()
        student = next((s for s in students if s.get('rollNo') == barcode), None)
        
        if student:
            return jsonify(student), 200
        else:
            return jsonify({'error': f'No student found with Roll Number: {barcode}'}), 404

    except Exception as e:
        print(f"Error in search_barcode: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/check_duplicate', methods=['POST'])
def check_duplicate():
    try:
        data = request.json
        roll_no = data.get('rollNo')
        reg_no = data.get('regNo')
        
        if not roll_no:
            return jsonify({'error': 'Roll number is required'}), 400

        students = load_students()
        
        # Check for duplicate roll number
        if any(s.get('rollNo') == roll_no for s in students):
            return jsonify({
                'isDuplicate': True,
                'message': f'Student with Roll Number {roll_no} already exists!'
            }), 200

        # Check for duplicate registration number if provided
        if reg_no and any(s.get('regNo') == reg_no for s in students):
            return jsonify({
                'isDuplicate': True,
                'message': f'Student with Registration Number {reg_no} already exists!'
            }), 200

        return jsonify({'isDuplicate': False}), 200

    except Exception as e:
        print(f"Error in check_duplicate: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/submit', methods=['POST'])
def submit():
    try:
        # Check content type to handle both form data and JSON requests
        student_data = {}
        if request.content_type and 'application/json' in request.content_type:
            # JSON request
            data = request.json
            for key, value in data.items():
                student_data[key] = value
        else:
            # Form data request
            for key in request.form:
                student_data[key] = request.form.get(key)
        
        print("Received student data:", student_data)  # Debug print
        
        # Handle profile image upload
        if 'profileImage' in request.files:
            profile_image = request.files['profileImage']
            if profile_image and profile_image.filename:
                # Only process if roll number is provided
                if 'rollNo' in student_data and student_data['rollNo']:
                    # Get file extension
                    filename = secure_filename(profile_image.filename)
                    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'jpg'
                    
                    # Create new filename using roll number
                    new_filename = f"{str(student_data['rollNo']).strip().lower()}.{file_ext}"
                    
                    # Ensure upload directory exists
                    upload_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
                    if not os.path.exists(upload_path):
                        os.makedirs(upload_path)
                    
                    # Save file with compression
                    # Save main as WEBP
                    new_filename = f"{str(student_data['rollNo']).strip().lower()}.webp"
                    file_path = os.path.join(upload_path, new_filename)
                    try:
                        profile_image.stream.seek(0)
                        with Image.open(profile_image.stream) as img:
                            if img.mode in ("P", "RGBA"):
                                img = img.convert("RGB")
                            max_size = (800, 800)
                            img.thumbnail(max_size, Image.LANCZOS)
                            img.save(file_path, format='WEBP', quality=80, method=6)
                            # Generate WebP thumbnail
                            try:
                                thumb_dir = os.path.join(upload_path, 'thumbs')
                                if not os.path.exists(thumb_dir):
                                    os.makedirs(thumb_dir)
                                thumb_path = os.path.join(thumb_dir, f"{str(student_data['rollNo']).strip().lower()}_thumb.webp")
                                thumb = img.copy()
                                thumb.thumbnail((256, 256), Image.LANCZOS)
                                thumb.save(thumb_path, format='WEBP', quality=75, method=6)
                            except Exception:
                                pass
                    except Exception:
                        profile_image.stream.seek(0)
                        profile_image.save(file_path)
                    
                    # Store the relative path in student data
                    student_data['profileImage'] = f"uploads/{new_filename}"
                    print(f"Set profile image path to: {student_data['profileImage']}")
        
        # Remove None and empty string values
        student_data = {k: v for k, v in student_data.items() if v is not None and v != ''}
        
        print("Processed student data:", student_data)  # Debug print

        # Validate required fields
        if not student_data.get('rollNo'):
            return jsonify({'error': 'Roll number is required'}), 400

        try:
            # Check for duplicates before saving
            students = load_students()
            
            # First check if document with this roll number already exists
            if any(s.get('rollNo') == student_data['rollNo'] for s in students):
                return jsonify({
                    'error': f'Student with Roll Number {student_data["rollNo"]} already exists!'
                }), 409

            # Check for duplicate registration number if provided
            if student_data.get('regNo') and any(s.get('regNo') == student_data['regNo'] for s in students):
                return jsonify({
                    'error': f'Student with Registration Number {student_data["regNo"]} already exists!'
                }), 409

            # Add timestamp
            student_data['createdAt'] = datetime.now().isoformat()
            
            # Save to JSON
            students.append(student_data)
            save_students(students)
            print(f"New student data saved for roll number: {student_data['rollNo']}")
            
            return jsonify({
                'success': True,
                'message': 'Student data saved successfully!',
                'studentId': student_data['rollNo']
            }), 200

        except Exception as e:
            print(f"Error saving student data: {str(e)}")
            return jsonify({'error': f'Database error: {str(e)}'}), 500

    except Exception as e:
        print(f"Error in submit: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/manage')
def manage_students():
    try:
        print("Starting manage_students route...")
        students = load_students()
        
        # Convert to list of dictionaries and organize by class
        students_list = []
        students_by_class = {}
        
        student_count = 0
        for student in students:
            student_data = student
            student_data['id'] = student_data['rollNo']
            students_list.append(student_data)
            student_count += 1
            
            # Organize students by class
            class_section = student_data.get('classSection', 'Unassigned')
            if class_section not in students_by_class:
                students_by_class[class_section] = []
            students_by_class[class_section].append(student_data)
        
        print(f"Found {student_count} students")
        print(f"Classes found: {list(students_by_class.keys())}")
        
        # Sort students in each class by roll number
        for class_section in students_by_class:
            students_by_class[class_section].sort(key=lambda x: x.get('rollNo', ''))
        
        # Sort the main list by roll number
        students_list.sort(key=lambda x: x.get('rollNo', ''))
        
        print("Rendering template with data...")
        return render_template('student_management.html', 
                            students=students_list,
                            students_by_class=students_by_class,
                            has_students=student_count > 0)
    except Exception as e:
        print(f"Error in manage_students: {str(e)}")
        # Return empty lists instead of error
        return render_template('student_management.html',
                            students=[],
                            students_by_class={},
                            has_students=False)

@app.route('/search_students', methods=['POST'])
def search_students():
    try:
        data = request.json
        query = data.get('query', '').strip().lower()
        filter_type = data.get('filter', 'all')
        
        if not query:
            return jsonify({'students': []}), 200

        students = load_students()
        filtered_students = []
        
        for student in students:
            # Get student data with default values
            student_name = student.get('studentName', '').lower()
            roll_no = student.get('rollNo', '').lower()
            class_section = student.get('classSection', '').lower()
            
            # Apply search based on filter type
            match_found = False
            
            if filter_type == 'all':
                # Search in all relevant fields
                match_found = (query in student_name or 
                             query in roll_no or 
                             query in class_section)
            elif filter_type == 'name':
                match_found = query in student_name
            elif filter_type == 'roll':
                match_found = query in roll_no
            elif filter_type == 'class':
                match_found = query in class_section
            
            if match_found:
                # Return all student data
                filtered_students.append(student)
        
        # Sort results by roll number
        filtered_students.sort(key=lambda x: x.get('rollNo', ''))
        
        # Limit to 20 results
        filtered_students = filtered_students[:20]
        
        return jsonify({'students': filtered_students}), 200
        
    except Exception as e:
        print(f"Error in search_students: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/advanced_search', methods=['POST'])
def advanced_search():
    try:
        criteria = request.json
        students = load_students()
        filtered_students = []
        
        for student in students:
            match = True
            
            # Name criteria
            if criteria.get('name') and criteria['name'].strip():
                if criteria['name'].lower() not in student.get('studentName', '').lower():
                    match = False
            
            # Roll number criteria
            if criteria.get('rollNo') and criteria['rollNo'].strip():
                if criteria['rollNo'].lower() not in student.get('rollNo', '').lower():
                    match = False
            
            # Class section criteria
            if criteria.get('classSection') and criteria['classSection'].strip():
                if criteria['classSection'].lower() not in student.get('classSection', '').lower():
                    match = False
            
            # Gender criteria
            if criteria.get('gender') and criteria['gender'].strip():
                if criteria['gender'].lower() != student.get('gender', '').lower():
                    match = False
            
            # Blood group criteria
            if criteria.get('bloodGroup') and criteria['bloodGroup'].strip():
                if criteria['bloodGroup'].lower() != student.get('bloodGroup', '').lower():
                    match = False
            
            # Category criteria
            if criteria.get('category') and criteria['category'].strip():
                if criteria['category'].lower() not in student.get('category', '').lower():
                    match = False
            
            # Address criteria
            if criteria.get('address') and criteria['address'].strip():
                if criteria['address'].lower() not in student.get('address', '').lower():
                    match = False
            
            if match:
                # Return all student data
                filtered_students.append(student)
        
        # Sort results by roll number
        filtered_students.sort(key=lambda x: x.get('rollNo', ''))
        
        # Limit to 30 results
        filtered_students = filtered_students[:30]
        
        return jsonify({'students': filtered_students}), 200
        
    except Exception as e:
        print(f"Error in advanced_search: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/get_student_details/<student_id>')
def get_student_details(student_id):
    try:
        students = load_students()
        student = next((s for s in students if s.get('rollNo') == student_id), None)
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
            
        return jsonify(student), 200
    except Exception as e:
        print(f"Error in get_student_details: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/edit_student/<student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    try:
        # For GET requests, display the edit form
        if request.method == 'GET':
            students = load_students()
            student = next((s for s in students if s.get('rollNo') == student_id), None)
            
            if student:
                return render_template('student_update.html', student=student)
            else:
                flash(f'Student with ID {student_id} not found', 'error')
                return redirect(url_for('manage_students'))
                
        # For POST requests, process the form submission
        elif request.method == 'POST':
            # Get data from form
            updated_data = {}
            for key in request.form:
                updated_data[key] = request.form.get(key)
                
            print(f"Edit student form data received for {student_id}: {updated_data}")
            print(f"Files in request: {list(request.files.keys())}")
                
            # Handle profile image upload
            if 'profileImage' in request.files:
                profile_image = request.files['profileImage']
                print(f"Profile image file: {profile_image}, filename: {profile_image.filename}")
                if profile_image and profile_image.filename:
                    # Get file extension
                    filename = secure_filename(profile_image.filename)
                    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'jpg'
                    
                    # Create new filename using student ID (roll number)
                    new_filename = f"{str(student_id).strip().lower()}.{file_ext}"
                    
                    # Ensure upload directory exists
                    upload_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
                    if not os.path.exists(upload_path):
                        os.makedirs(upload_path)
                    
                    # Save file with compression
                    # Save main as WEBP
                    new_filename = f"{str(student_id).strip().lower()}.webp"
                    file_path = os.path.join(upload_path, new_filename)
                    print(f"Saving profile image to: {file_path}")
                    try:
                        profile_image.stream.seek(0)
                        with Image.open(profile_image.stream) as img:
                            if img.mode in ("P", "RGBA"):
                                img = img.convert("RGB")
                            max_size = (800, 800)
                            img.thumbnail(max_size, Image.LANCZOS)
                            img.save(file_path, format='WEBP', quality=80, method=6)
                            # Generate WebP thumbnail
                            try:
                                thumb_dir = os.path.join(upload_path, 'thumbs')
                                if not os.path.exists(thumb_dir):
                                    os.makedirs(thumb_dir)
                                thumb_path = os.path.join(thumb_dir, f"{str(student_id).strip().lower()}_thumb.webp")
                                thumb = img.copy()
                                thumb.thumbnail((256, 256), Image.LANCZOS)
                                thumb.save(thumb_path, format='WEBP', quality=75, method=6)
                            except Exception:
                                pass
                    except Exception:
                        profile_image.stream.seek(0)
                        profile_image.save(file_path)
                    print(f"Saved profile image to {file_path}")
                    
                    # Store the relative path in student data
                    updated_data['profileImage'] = f"uploads/{new_filename}"
                    print(f"Set profile image path to: {updated_data['profileImage']}")
                
            # Remove None and empty string values
            updated_data = {k: v for k, v in updated_data.items() if v is not None and v != ''}
            
            # Update the student in the JSON file
            students = load_students()
            
            # Find the student by ID (roll number)
            student_index = next((i for i, s in enumerate(students) if s.get('rollNo') == student_id), None)
            
            if student_index is not None:
                # Update the student data
                students[student_index].update(updated_data)
                
                # Save the updated students list
                save_students(students)
                
                flash('Student data updated successfully!', 'success')
            else:
                flash(f'Student with ID {student_id} not found', 'error')
                
            return redirect(url_for('manage_students'))

    except Exception as e:
        print(f"Error in edit_student: {str(e)}")
        flash(f'Error updating student: {str(e)}', 'error')
        return redirect(url_for('manage_students'))

@app.route('/delete_student/<student_id>', methods=['DELETE'])
def delete_student(student_id):
    try:
        students = load_students()
        student = next((s for s in students if s.get('rollNo') == student_id), None)
        
        if not student:
            return jsonify({
                'success': False,
                'error': 'Student not found'
            }), 404
            
        # Remove the student from the list
        students = [s for s in students if s.get('rollNo') != student_id]
        
        # Save updated students
        save_students(students)
        
        return jsonify({
            'success': True,
            'message': 'Student deleted successfully',
            'deletedId': student_id
        })
            
    except Exception as e:
        print(f"Error deleting student: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Only .xlsx files are allowed'}), 400
        
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'])
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Read Excel file
            df = pd.read_excel(filepath,dtype={'Student Contact': str, 'Parent No': str})
            if df.empty:
                os.remove(filepath)
                return jsonify({'error': 'Excel file is empty'}), 400
            
            print("Excel columns:", df.columns.tolist())
            
            # Column mapping for Excel to database fields
            column_mapping = {
                # Basic Info
                'Roll No': 'rollNo',
                'Reg No': 'regNo',
                'Section': 'classSection',
                'Student Name': 'studentName',
                'Father Name': 'fatherName',
                'Gender': 'gender',
                'DOB': 'dob',
                
                # Contact Info
                'Email': 'email',
                'Address': 'address',
                'Student Contact': 'studentContact',
                'Parent No': 'studentContact',
                
                # Personal Info
                'Aadhar No': 'aadharNo',
                'Blood Group': 'bloodGroup',
                'State': 'state',
                'District': 'district',
                'Religion': 'religion',
                'Category': 'category',
                'Caste': 'caste',
                'Income': 'income',
                'Category Applied': 'categoryApplied',
                
                # PUC Details
                'PUC/Equivalent Roll No': 'pucRollNo',
                'PUC/Equivalent Year of Completion': 'pucYear',
                'PUC/Equivalent Institute Name': 'pucInstitute',
                'PUC/Equivalent Total Marks': 'pucTotalMarks',
                'PUC/Equivalent Obtained Marks': 'pucObtainedMarks',
                'PUC/Equivalent Percentage/CGPA': 'pucPercentage',
                
                # Program Details
                'Program Name': 'programName',
                'Discipline 1': 'discipline1',
                'Language2': 'lang2',
                
                # ABC ID Details
                'ABC ID': 'abcId'
            }
            
            success_count = 0
            error_count = 0
            error_messages = []

            df['Student Contact'] = df['Student Contact'].str.replace(r'\.0$', '', regex=True)
            df['Parent No'] = df['Parent No'].str.replace(r'\.0$', '', regex=True)

            
            # Load existing students
            students = load_students()
            print(f"Loaded {len(students)} existing students")
            
            # Process each row
            for index, row in df.iterrows():
                try:
                    # Initialize student data with default values
                    student_data = {
                        'classSection': 'N/A',
                        'collegeName': 'N/A',
                        'programName': 'N/A',
                        'discipline1': 'N/A',
                        'lang2': 'N/A',
                        'studentContact': 'N/A',
                        'email': 'N/A',
                        'pucRollNo': 'N/A',
                        'pucYear': 'N/A',
                        'pucInstitute': 'N/A',
                        'pucTotalMarks': '0',
                        'pucObtainedMarks': '0',
                        'pucPercentage': '0',
                        'abcId': 'N/A',
                        'abcIdCollected': 'No'
                    }
                    
                    # Map Excel columns to student fields
                    for col in df.columns:
                        if col in column_mapping:
                            value = row[col]
                            if pd.notna(value):  # Check if value is not NaN
                                value = str(value).strip()
                                if value.upper() not in ['NA', 'N/A', 'NULL', 'NONE', '', '-', 'EMPTY']:
                                    db_field = column_mapping.get(col)
                                    if db_field:
                                        student_data[db_field] = value
                    
                    # Validate required fields
                    if not student_data.get('rollNo'):
                        error_messages.append(f"Row {index + 2}: Missing Roll Number")
                        error_count += 1
                        continue
                    
                    # Check if student exists
                    student_exists = False
                    for i, student in enumerate(students):
                        if student['rollNo'] == student_data['rollNo']:
                            # Update existing student
                            students[i] = {**student, **student_data}
                            student_exists = True
                            print(f"Updated existing student: {student_data['rollNo']}")
                            break
                    
                    if not student_exists:
                        # Add new student
                        student_data['createdAt'] = datetime.now().isoformat()
                        students.append(student_data)
                        print(f"Added new student: {student_data['rollNo']}")
                    
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1
                    print(f"Error in row {index + 2}: {str(e)}")
            
            # Save all students at once
            save_students(students)
            print(f"Saved {len(students)} total students")
            
            # Clean up
            os.remove(filepath)
            
            # Return appropriate message based on operation type
            return jsonify({
                'success': True,
                'message': f'Processed {len(df)} rows with {success_count} successes and {error_count} issues.',
                'stats': {
                    'total': len(df),
                    'success': success_count,
                    'errors': error_count
                },
                'error_messages': error_messages
            })
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': f'Error processing Excel file: {str(e)}'}), 500
            
    except Exception as e:
        return jsonify({'error': f'Error handling file upload: {str(e)}'}), 500

@app.route('/delete_class/<class_section>', methods=['DELETE'])
def delete_class(class_section):
    try:
        if not class_section:
            return jsonify({
                'success': False,
                'error': 'Class section is required'
            }), 400
            
        students = load_students()
        students_to_delete = [s for s in students if s.get('classSection') == class_section]
        if not students_to_delete:
            return jsonify({
                'success': False,
                'error': f'No students found in class {class_section}'
            }), 404
            
        deleted_count = 0
        try:
            # Filter out students to be deleted
            students_filtered = [s for s in students if s.get('classSection') != class_section]
            
            # Save updated students
            save_students(students_filtered)
            
            deleted_count = len(students_to_delete)
            return jsonify({
                'success': True,
                'message': f'Successfully deleted {deleted_count} students from class {class_section}',
                'deletedCount': deleted_count
            })
        except Exception as e:
            print(f"Error in batch delete: {str(e)}")
            return jsonify({
                'success': False,
                'error': f'Database error while deleting class: {str(e)}'
            }), 500
            
    except Exception as e:
        print(f"Error deleting class: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/edit_class/<class_section>', methods=['PUT'])
def edit_class(class_section):
    try:
        if not class_section:
            return jsonify({
                'success': False,
                'error': 'Class section is required'
            }), 400
            
        data = request.get_json()
        new_class_name = data.get('newClassName')
        
        if not new_class_name:
            return jsonify({
                'success': False,
                'error': 'New class name is required'
            }), 400
            
        students = load_students()
        students_to_update = [s for s in students if s.get('classSection') == class_section]
        if not students_to_update:
            return jsonify({
                'success': False,
                'error': f'No students found in class {class_section}'
            }), 404
            
        updated_count = 0
        try:
            # Update all students in a batch
            for i, student in enumerate(students_to_update):
                students[i] = {**student, **{'classSection': new_class_name}}
            
            # Save updated students
            save_students(students)
            
            updated_count = len(students_to_update)
            return jsonify({
                'success': True,
                'message': f'Successfully renamed class from {class_section} to {new_class_name}',
                'updatedCount': updated_count
            })
        except Exception as e:
            print(f"Error in batch update: {str(e)}")
            return jsonify({
                'success': False,
                'error': f'Database error while renaming class: {str(e)}'
            }), 500
            
    except Exception as e:
        print(f"Error renaming class: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/refresh_data', methods=['GET'])
def refresh_data():
    try:
        # Return the latest data
        students = load_students()
        student_list = []
        for student in students:
            student_data = student
            student_data['id'] = student['rollNo']
            student_list.append(student_data)
            
        return jsonify({
            'success': True,
            'data': student_list,
            'timestamp': datetime.now().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/verify_data_consistency', methods=['POST'])
def verify_data_consistency():
    try:
        data = request.json
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'error': 'Student ID is required'}), 400
            
        # Get the latest data from JSON
        students = load_students()
        student = next((s for s in students if s.get('rollNo') == student_id), None)
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
            
        return jsonify({
            'success': True,
            'data': student,
            'last_updated': student.get('updatedAt', None)
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/compare_student_data', methods=['POST'])
def compare_student_data():
    try:
        data = request.json
        excel_data = data.get('excelData', {})
        roll_no = excel_data.get('rollNo')
        
        if not roll_no:
            return jsonify({'error': 'Roll number is required'}), 400
            
        # Get student from JSON
        students = load_students()
        student = next((s for s in students if s.get('rollNo') == roll_no), None)
        
        if not student:
            return jsonify({'error': 'Student not found in JSON'}), 404
            
        # Compare fields and find differences
        differences = {}
        for field, excel_value in excel_data.items():
            if field in student:
                db_value = str(student[field]).strip()
                excel_value = str(excel_value).strip()
                
                # Skip empty or matching values
                if excel_value and excel_value != 'Empty' and excel_value != db_value:
                    differences[field] = {
                        'database': db_value,
                        'excel': excel_value,
                        'field_name': field.replace('_', ' ').title()  # Format field name for display
                    }
        
        # Return comparison results with student info
        return jsonify({
            'hasDifferences': len(differences) > 0,
            'differences': differences,
            'student_info': {
                'rollNo': roll_no,
                'studentName': student.get('studentName', 'N/A'),
                'classSection': student.get('classSection', 'N/A')
            },
            'message': f'Found {len(differences)} differences for {student.get("studentName", "Student")}' if differences else 'No differences found'
        })
        
    except Exception as e:
        print(f"Error in compare_student_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/update_student_data', methods=['POST'])
def update_student_data():
    try:
        data = request.json
        roll_no = data.get('rollNo')
        updates = data.get('updates', {})
        
        if not roll_no or not updates:
            return jsonify({'error': 'Roll number and updates are required'}), 400
            
        # Get student data
        students = load_students()
        student = next((s for s in students if s.get('rollNo') == roll_no), None)
        
        if not student:
            return jsonify({'error': 'Student not found'}), 404
            
        # Update only the specified fields
        updated_student = {**student, **updates, 'lastUpdated': datetime.now().isoformat()}
        
        # Update the student data
        for i, s in enumerate(students):
            if s['rollNo'] == roll_no:
                students[i] = updated_student
                break
        
        # Save updated students
        save_students(students)
        
        return jsonify({
            'success': True,
            'message': f'Successfully updated {len(updates)} field(s)',
            'updatedFields': list(updates.keys())
        })
        
    except Exception as e:
        print(f"Error in update_student_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_template')
def download_template():
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Data Template"

        # Define headers
        headers = [
            'Roll No', 'Reg No', 'Section', 'Student Name', 'Father Name', 'Gender', 'DOB',
            'Email', 'Address', 'Student Contact', 'Parent No', 'Aadhar No', 'Blood Group',
            'State', 'District', 'Religion', 'Category', 'Caste', 'Income', 'Category Applied',
            'PUC/Equivalent Roll No', 'PUC/Equivalent Year of Completion', 'PUC/Equivalent Institute Name',
            'PUC/Equivalent Total Marks', 'PUC/Equivalent Obtained Marks', 'PUC/Equivalent Percentage/CGPA',
            'Program Name', 'Discipline 1', 'Language2', 'ABC ID'
        ]

        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Add data validation for specific columns
        gender_validation = DataValidation(
            type="list",
            formula1='"Male,Female,Other"',
            allow_blank=True
        )
        ws.add_data_validation(gender_validation)
        gender_validation.add(f"F2:F1000")  # Apply to Gender column

        # Save to BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='student_template.xlsx'
        )

    except Exception as e:
        print(f"Error creating template: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/debug/profile-images')
def debug_profile_images():
    """Debug route to check profile image paths"""
    try:
        # Get all students
        students = load_students()
        
        # Extract profile image information
        image_info = []
        for student in students:
            info = {
                'rollNo': student.get('rollNo', 'N/A'),
                'studentName': student.get('studentName', 'N/A'),
                'hasProfileImage': 'profileImage' in student,
                'profileImagePath': student.get('profileImage', 'None')
            }
            
            # Check if file exists on disk
            if info['hasProfileImage'] and info['profileImagePath']:
                full_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', student['profileImage'])
                info['fileExists'] = os.path.exists(full_path)
                info['fullPath'] = full_path
            else:
                info['fileExists'] = False
                info['fullPath'] = 'N/A'
                
            image_info.append(info)
            
        # Check upload folder
        upload_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
        files_in_folder = []
        if os.path.exists(upload_folder):
            files_in_folder = os.listdir(upload_folder)
            
        return jsonify({
            'students_with_images': sum(1 for s in image_info if s['hasProfileImage']),
            'total_students': len(image_info),
            'image_info': image_info,
            'upload_folder': upload_folder,
            'files_in_upload_folder': files_in_folder
        }), 200
        
    except Exception as e:
        print(f"Error in debug_profile_images: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/bulk_image_upload', methods=['POST'])
def bulk_image_upload():
    """Handle bulk upload of student profile images"""
    try:
        # Check if any files were uploaded
        if 'images' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
            
        uploaded_files = request.files.getlist('images')
        if len(uploaded_files) == 0:
            return jsonify({'error': 'No files selected'}), 400
            
        # Load all students to match roll numbers
        students = load_students()
        student_roll_numbers = {student.get('rollNo', '').lower(): student for student in students}
        
        # Set up upload folder
        upload_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
        if not os.path.exists(upload_path):
            os.makedirs(upload_path)
            
        results = {
            'success': [],
            'errors': []
        }
        
        # Process each file
        for file in uploaded_files:
            try:
                # Get the original filename and extract roll number
                original_filename = secure_filename(file.filename)
                
                # Try different methods to extract roll number:
                # 1. Use the filename without extension as roll number
                filename_without_ext = os.path.splitext(original_filename)[0].lower()
                
                # 2. Check if file matches an existing roll number
                matching_roll = None
                if filename_without_ext in student_roll_numbers:
                    matching_roll = filename_without_ext
                
                # If no match found, skip this file
                if not matching_roll:
                    results['errors'].append({
                        'filename': original_filename,
                        'error': f'No student found with roll number matching filename: {filename_without_ext}'
                    })
                    continue
                    
                # Get the student and determine file extension
                student = student_roll_numbers[matching_roll]
                file_ext = os.path.splitext(original_filename)[1].lower()
                if not file_ext:
                    file_ext = '.jpg'  # Default to jpg if no extension
                
                # Create the new filename
                new_filename = f"{matching_roll}{file_ext}"
                file_path = os.path.join(upload_path, new_filename)
                
                # Save the file with compression
                try:
                    file.stream.seek(0)
                    with Image.open(file.stream) as img:
                        if img.mode in ("P", "RGBA"):
                            img = img.convert("RGB")
                        max_size = (800, 800)
                        img.thumbnail(max_size, Image.LANCZOS)
                        # Save main as WEBP
                        new_filename = f"{matching_roll}.webp"
                        file_path = os.path.join(upload_path, new_filename)
                        img.save(file_path, format='WEBP', quality=80, method=6)
                        # Generate WebP thumbnail
                        try:
                            thumb_dir = os.path.join(upload_path, 'thumbs')
                            if not os.path.exists(thumb_dir):
                                os.makedirs(thumb_dir)
                            thumb_path = os.path.join(thumb_dir, f"{matching_roll}_thumb.webp")
                            thumb = img.copy()
                            thumb.thumbnail((256, 256), Image.LANCZOS)
                            thumb.save(thumb_path, format='WEBP', quality=75, method=6)
                        except Exception:
                            pass
                except Exception:
                    file.stream.seek(0)
                    file.save(file_path)
                
                # Update the student record with the image path
                image_rel_path = f"uploads/{new_filename}"
                
                # Find the student in the original list and update
                for s in students:
                    if s.get('rollNo', '').lower() == matching_roll:
                        s['profileImage'] = image_rel_path
                        break
                
                # Add to success results
                results['success'].append({
                    'original_filename': original_filename,
                    'roll_number': matching_roll,
                    'new_filename': new_filename,
                    'image_path': image_rel_path
                })
                
            except Exception as e:
                results['errors'].append({
                    'filename': original_filename if 'original_filename' in locals() else 'unknown',
                    'error': str(e)
                })
        
        # Save updated student data
        save_students(students)
        
        # Return results
        return jsonify({
            'success': True,
            'message': f"Successfully processed {len(results['success'])} images with {len(results['errors'])} errors",
            'results': results
        })
        
    except Exception as e:
        print(f"Error in bulk_image_upload: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True) 
